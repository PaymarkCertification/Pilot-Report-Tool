import PySimpleGUI as sg
import typing
import os
import configparser
import logging
import datetime
import glob
import subprocess
import sys
import pandas as pd
import numpy as np
import xlsxwriter
from openpyxl.styles import PatternFill
__version__ = '0.0.3'



# ----- var ----- #
# Initialize our containers
filename = ''
tableList=[]
frameDataForProcessing = pd.DataFrame()


# ----- Logging Method ----- #
class Handler(logging.StreamHandler):

    def __init__(self):
        logging.StreamHandler.__init__(self)

    # def emit(self, record):
    #     global buffer
    #     record = f'{record.name}, [{record.levelname}], {record.message}'
    #     buffer = f'{buffer}\n{record}'.strip()
    #     window['-output-'].update(value=buffer)

    @staticmethod
    def load_logging(loggingLevel: int = logging.INFO) -> logging:
        return logging.basicConfig(
            level=loggingLevel, 
            format='%(asctime)s %(levelname)s:%(message)s', 
            handlers=[
                logging.StreamHandler()
            ]
        )

# ----- exception handling func ----- #
# decorator 
def error_handler(func: typing.Callable):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except:
            print(f'function {func.__name__} encountered an exception: {sys.exc_info()[0], sys.exc_info()[1]}') # Outputs to console
            logging.info(f'{func.__name__, sys.exc_info()[0], sys.exc_info()[1]}')
    return wrapper

# ----- GUI Methods ---- #
def load_config() -> str:
    """Directory for RPT files. Set accordingly."""
    config = configparser.ConfigParser()
    config.read(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'Pilot_Report_Config.ini'))
    path = config.get('path', 'rpt_file')
    return path

@error_handler
def lastweek(string: int = 1) -> typing.Union[str, datetime.datetime]:
    """
    @param string: Toggles return type
    Returns last 7 days from today either as string or date"""
    today = datetime.datetime.now()
    week_ago = today - datetime.timedelta(days=7)
    if string:
        return week_ago.strftime('%d/%m/%Y') # returns string type
    else:
        return week_ago # returns date object

@error_handler
def last_days() -> datetime.datetime:
    date = values['datePick']
    logging.info(f'date returned from widget: {date}')
    selected_date = datetime.datetime.strptime(date,'%d/%m/%Y')
    week = selected_date - datetime.timedelta(days=7)    
    return week

@error_handler
def set_excel_format(excel_name: str):
    import openpyxl
    from openpyxl.styles import PatternFill
    # Const
    RED_FORMAT = 'FFFF0000'
    GREEN_FORMAT = '008000'
    currentstan = 0
    laststan = 0
    workbook = openpyxl.load_workbook(excel_name)
    ws = workbook.active
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows(min_row=2, min_col=6,max_col=6):
            for cell in row:
                if cell.value is None or cell.value == '\n':
                    continue
                if cell.value:
                    laststan = currentstan
                    stan = int(cell.value)
                    currentstan = stan
                    if currentstan != laststan + 1 and row != 0:
                        ws[rower].fill = PatternFill(start_color=RED_FORMAT,
                                                    end_color=RED_FORMAT,
                                                    fill_type='solid')
                    else:
                        ws[rower].fill = PatternFill(start_color=GREEN_FORMAT,
                                                    end_color=GREEN_FORMAT,
                                                    fill_type='solid')
                else:
                    pass
    workbook.save(excel_name)
    workbook.close()

@error_handler
def process_rpt_files():
    def concatentate_csvs(list_of: list) -> pd.DataFrame:
        li = []
        logging.info(' Attempting to concatenate')
        
        try:
            
            for filename in list_of:
                df = pd.read_csv(filename, index_col=None, header=0, sep='|')
                li.append(df)
            dframe = pd.concat(li, axis=0, ignore_index=True)
            dframe.replace(np.nan, '', regex=True)
            dframe['Card Number'] = dframe['Card Number'].convert_dtypes(convert_integer=False, convert_boolean=False)
            dframe['Terminal'] = dframe['Terminal'].apply(str)
            global frameDataForProcessing; frameDataForProcessing = dframe
            tableframe = dframe.groupby(['Terminal','Card Number'], as_index=False).last()
            tableframe = tableframe[tableframe['Card Number'].str.contains('^[A-Za-z]')]
            
            
            global tableList; tableList.append(tableframe[['Terminal','Card Number']].values.tolist())
            window.FindElement('-Table-').Update(tableframe[['Terminal','Card Number']].values.tolist())
        except ValueError as e:
            print(f'Err: Unable to perform function concatenate csvs. Check date range or RPT files.')
            # print(f' {sys.exc_info()[0], sys.exc_info()[1]}')
            logging.info(f' {sys.exc_info()[0], sys.exc_info()[1]}')
        
           
    try:
        os.chdir(load_config())
    except FileNotFoundError:
        print(f'The system cannot find the file specified in {load_config.__name__}. '
        f'\nPlease check Pilot_Report_Config.ini is configured correctly.'
        f'\nCurrent Working Dir: {load_config()}')
        logging.info(' Could not load file')

    print(f'Harvesting data from {values["datePick"]}')
    fileDirectory = [f for f in glob.glob("*.txt")]
    formattedDate=[]
    try:
        for rptfile in fileDirectory:
            new = rptfile[6:-4] # strips the chars to retrieve date from file name
            formattedDate.append(new.format(datetime.datetime.strptime(new, '%y%m%d')))  
    except ValueError as e:
        print(f'{process_rpt_files.__name__}: Value Error. Empty list.')


    dateObjects = []
    for date in formattedDate:
        convertedDateObject = datetime.datetime.strptime(date, '%y%m%d') # converts stripped file name into date object
        logging.info(f' Convert str "{date}" to date object "{convertedDateObject}"')

        if convertedDateObject >= datetime.datetime.strptime(values['datePick'], '%d/%m/%Y') and convertedDateObject <= datetime.datetime.strptime(values['datePick'], '%d/%m/%Y')+ datetime.timedelta(7): 
            dateObjects.append(convertedDateObject.strftime('%y%m%d')) # adds the date of the file to list if they are within the last 7 days
            logging.info(f' {convertedDateObject} appended to list')

    list_of_files = []
    for date in dateObjects:
        for fileItem in fileDirectory:
            if date in fileItem:
                list_of_files.append(fileItem) # Returns the filename for processing if the dates are in the filename
    if not list_of_files:
        print('No files found for selected date range. Select another date.\n')
    elif list_of_files != '':
        print()
        print('Processing the following files:\n', list_of_files, '\n')
        concatentate_csvs(list_of_files)
    else:
        print("No files found for selected date range. Select another date.")

@error_handler
def grab_user_selected_id(selectionList: list) -> list:
    if not selectionList:
        print("\n")
    else:
        users_list =[]
        for i in values['-Table-']:
            users_list.append(tableList[0][i])
        user_id=[item[0] for item in users_list]
        print(f'Terminal ID(s) selected for report:\n{user_id}\n')    
        return user_id



@error_handler
def process_report(user_id_list: list):
    """
    @param user_id_list: Returned from table widget
    """
    if not user_id_list:
        print("No ID(s) selected for processing")
    else:
        foldername = sg.PopupGetFolder('Select folder', no_window=True)
        date_time = datetime.datetime.today().strftime("%d-%m-%Y %H%M%S")
        global filename; filename = f'{foldername}/Pilot_Report_{date_time}.xlsx'
        print('File: ',filename) 
        writer = pd.ExcelWriter(f'{filename}', engine='openpyxl')
        global frameDataForProcessing
        for userID in user_id_list:
            # index_len = str(len(frameDataForProcessing["Terminal"].index == userID)) # FuturesError warning raised when comparing np to py scalar
            frameDataForProcessing.loc[frameDataForProcessing['Terminal'].astype(str) == userID]\
                .to_excel(writer, index=False, sheet_name=userID)  # set sheet name and writes sheet data with terminal ID from dataframe
            worksheet = writer.book[userID]

            print(f'processing {userID} data')

            # column formatting on DF
            dims = {}
            for row in worksheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                worksheet.column_dimensions[col].width = value

            # TODO: Remove setter loop.
            # column setter for xlsxwriter engine
            # for column in frameDataForProcessing:
            #     column_width = max(frameDataForProcessing[column].astype(str).map(len).max(), len(column))
            #     col_idx = frameDataForProcessing.columns.get_loc(column)
            #     writer.sheets[str(userID)].set_column(col_idx, col_idx, column_width)
        writer.save()
        print('Applying formatting')
        set_excel_format(filename)
        print('Excel saved')

# ----- Column Definition ----- #
colx = [
    [sg.Text("Date:"), sg.Input(lastweek(), key='datePick', size=(22, 1)),sg.CalendarButton("Select Date", key='date', disabled=False, format='%d/%m/%Y')],
    # [sg.CalendarButton("Select Start Date", key='date', disabled=False, format='%d/%m/%Y')],
    [sg.Radio('Translate Data', group_id='RADIO1', default=False), sg.Radio('Raw Data', group_id='RADIO1', default=True)],
    [sg.Button('Harvest'), sg.Button('Collect'), sg.Button('Clear Log'), sg.Button('path')]
]

coly = [
    [sg.Table(key='-Table-', headings=['Terminal ID', 'Software Version'],
     values=[['', '']], col_widths=500,
     num_rows=7, enable_events=True)]
]

colz = [sg.Output(size=(81, 9), key='-output-')]

# ----- Set Theme ----- #
# sg.theme('BlueMono')

# ----- Layout Management ----- #
layout = [
    [sg.Column(colx),sg.VerticalSeparator(pad=None), sg.Column(coly)],
    [sg.Output(size=(81, 9), key='-output-')]
]
window = sg.Window(f'Pilot Report - {__version__}', layout)


if __name__=='__main__':
    Handler.load_logging()
    
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        elif event == 'Harvest':
            window['-Table-'].update('')
            process_rpt_files()
        elif event == 'Clear Log':
            window.FindElement('-output-').Update('')
        elif event == 'path':
            print(load_config())
        elif event == 'Collect':
            process_report(grab_user_selected_id(values['-Table-']))
        logging.info(f' {event} event called')